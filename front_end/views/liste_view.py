from tkinter import ttk, messagebox, filedialog
import customtkinter as ctk
from PIL import Image
from typing import Optional
from front_end.widgets.tooltip import Tooltip
import config

# Colonnes : (clé JSON, en-tête affiché, largeur px)
COLUMNS = [
    ("nom",                  "Nom du projet",        160),
    ("client",               "Client",               100),
    ("type_etude",           "Type d'études",        130),
    ("type_contrat",         "Type de contrat",      150),
    ("type_activites",       "Type d'activités",     120),
    ("type_calculs",         "Type de calculs",      155),
    ("logiciels_utilises",   "Logiciels",            140),
    ("type_donnees_entrees", "Données d'entrées",    120),
    ("statut",               "Statut",               100),
    ("date_debut",           "Date de début",        110),
    ("date_fin",             "Date de fin",          110),
    ("livrable",             "Livrables",            155),
    ("collaborateur_referent","Collaborateur réf.",  130),
    ("mots_cles",            "Mots clés",            110),
    ("commentaire",          "Commentaire",          150),
    ("directory",            "Directory",            160),
]


class ListeView(ctk.CTkFrame):
    """Vue tableau avec barre d'outils (actualiser, supprimer, modifier,
    dupliquer, exporter Excel, ajouter)."""

    def __init__(self, parent, gestionnaire, app):
        super().__init__(parent, fg_color="#f0f4f8", corner_radius=0)
        self.gestionnaire = gestionnaire
        self.app = app
        self._build()

    # ── Construction ───────────────────────────────────────────────────────────

    def _build(self):
        self._build_toolbar()
        self._build_table()

    def _build_toolbar(self):
        toolbar = ctk.CTkFrame(self, fg_color="white", height=54, corner_radius=0)
        toolbar.pack(fill="x")
        toolbar.pack_propagate(False)

        # Séparateur bas
        ctk.CTkFrame(toolbar, height=1, fg_color="#e5e7eb").place(
            relx=0, rely=1.0, relwidth=1, anchor="sw"
        )

        def _ico(filename, size=20):
            return ctk.CTkImage(Image.open(config.get_base_path() / "assets" / filename), size=(size, size))

        base_kw = dict(height=36, corner_radius=6, width=42,
                       fg_color="#f3f4f6", text_color="#374151", text="")

        btn_refresh = ctk.CTkButton(toolbar, image=_ico("rafraichir.png"), command=self.refresh,
                                    hover_color="#e5e7eb", **base_kw)
        btn_refresh.pack(side="left", padx=(12, 2), pady=9)
        Tooltip(btn_refresh, "Actualiser")

        btn_delete = ctk.CTkButton(toolbar, image=_ico("delete.png"), command=self._supprimer,
                                   hover_color="#fecaca", **base_kw)
        btn_delete.pack(side="left", padx=2, pady=9)
        Tooltip(btn_delete, "Supprimer")

        btn_edit = ctk.CTkButton(toolbar, image=_ico("edite.png"), command=self._modifier,
                                 hover_color="#bfdbfe", **base_kw)
        btn_edit.pack(side="left", padx=2, pady=9)
        Tooltip(btn_edit, "Modifier")

        btn_dup = ctk.CTkButton(toolbar, image=_ico("dupliquer.png"), command=self._dupliquer,
                                hover_color="#bbf7d0", **base_kw)
        btn_dup.pack(side="left", padx=2, pady=9)
        Tooltip(btn_dup, "Dupliquer")

        btn_export = ctk.CTkButton(toolbar, image=_ico("export.png"), command=self._exporter_excel,
                                   hover_color="#fef3c7", **base_kw)
        btn_export.pack(side="left", padx=2, pady=9)
        Tooltip(btn_export, "Exporter en Excel")

        # Séparateur vertical
        ctk.CTkFrame(toolbar, width=1, fg_color="#e5e7eb").pack(
            side="left", fill="y", pady=12, padx=10
        )

        ctk.CTkButton(
            toolbar,
            text="＋  Nouveau projet",
            width=155,
            height=36,
            corner_radius=6,
            fg_color="#2563eb",
            hover_color="#1d4ed8",
            text_color="white",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self.app.show_formulaire_new,
        ).pack(side="left", padx=2, pady=9)

    def _build_table(self):
        wrapper = ctk.CTkFrame(self, fg_color="white", corner_radius=8)
        wrapper.pack(fill="both", expand=True, padx=15, pady=15)

        # Style ttk
        style = ttk.Style()
        style.theme_use("default")
        style.configure(
            "Proj.Treeview",
            background="white",
            foreground="#1f2937",
            fieldbackground="white",
            rowheight=30,
            font=("Segoe UI", 10),
            borderwidth=0,
        )
        style.configure(
            "Proj.Treeview.Heading",
            background="#f9fafb",
            foreground="#374151",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            borderwidth=0,
        )
        style.map(
            "Proj.Treeview",
            background=[("selected", "#dbeafe")],
            foreground=[("selected", "#1e40af")],
        )
        style.layout("Proj.Treeview", [("Proj.Treeview.treearea", {"sticky": "nswe"})])

        col_ids = [c[0] for c in COLUMNS]

        vsb = ttk.Scrollbar(wrapper, orient="vertical")
        hsb = ttk.Scrollbar(wrapper, orient="horizontal")

        self._tree = ttk.Treeview(
            wrapper,
            columns=col_ids,
            show="headings",
            style="Proj.Treeview",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            selectmode="browse",
        )
        vsb.configure(command=self._tree.yview)
        hsb.configure(command=self._tree.xview)

        for col_id, label, width in COLUMNS:
            self._tree.heading(col_id, text=label, anchor="w")
            self._tree.column(col_id, width=width, minwidth=60, anchor="w", stretch=False)

        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        wrapper.grid_rowconfigure(0, weight=1)
        wrapper.grid_columnconfigure(0, weight=1)

        # Double-clic → modifier
        self._tree.bind("<Double-1>", lambda _e: self._modifier())

    # ── Données ────────────────────────────────────────────────────────────────

    @staticmethod
    def _val(projet: dict, key: str) -> str:
        """Formatte une valeur pour l'affichage dans le tableau."""
        # Rétro-compatibilité mots_cles / Mots_cles
        if key == "mots_cles" and "mots_cles" not in projet:
            val = projet.get("Mots_cles", "")
        else:
            val = projet.get(key, "")
        if val is None:
            return ""
        if isinstance(val, list):
            return ", ".join(str(v) for v in val if v)
        return str(val)

    def refresh(self):
        """Recharge les données depuis le JSON et met à jour le tableau."""
        self.gestionnaire.projets = self.gestionnaire._charger_projets()
        self._tree.delete(*self._tree.get_children())
        for projet in self.gestionnaire.projets:
            values = tuple(self._val(projet, col_id) for col_id, _, _ in COLUMNS)
            self._tree.insert("", "end", iid=projet["id"], values=values)

    # ── Sélection ──────────────────────────────────────────────────────────────

    def _selected_projet(self) -> Optional[dict]:
        sel = self._tree.selection()
        if not sel:
            messagebox.showwarning(
                "Aucune sélection", "Veuillez sélectionner un projet dans le tableau."
            )
            return None
        return self.gestionnaire.trouver_projet(sel[0])

    # ── Actions barre d'outils ─────────────────────────────────────────────────

    def _verifier_appartenance(self, projet: dict, action: str) -> bool:
        """Affiche un message d'erreur si le projet n'appartient pas à l'utilisateur.
        Retourne True si l'action est autorisée, False sinon."""
        if self.gestionnaire._appartient_a_utilisateur(projet):
            return True
        messagebox.showerror(
            "Action refusée",
            f"Vous ne pouvez pas {action} ce projet.\n\n"
            f"Il appartient à « {projet.get('id_utilisateur')} ».",
        )
        return False

    def _supprimer(self):
        projet = self._selected_projet()
        if not projet or not self._verifier_appartenance(projet, "supprimer"):
            return
        if messagebox.askyesno(
            "Confirmer la suppression",
            f"Supprimer le projet « {projet['nom']} » ?\n\nCette action est irréversible.",
            icon="warning",
        ):
            ok, erreur = self.gestionnaire.supprimer_projet(projet["id"])
            if not ok:
                messagebox.showerror("Erreur", erreur)
            else:
                self.refresh()

    def _modifier(self):
        projet = self._selected_projet()
        if not projet or not self._verifier_appartenance(projet, "modifier"):
            return
        self.app.show_formulaire_edit(projet, mode="modifier")

    def _dupliquer(self):
        projet = self._selected_projet()
        if projet:
            self.app.show_formulaire_edit(projet, mode="dupliquer")

    def _exporter_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Module manquant",
                "Le module openpyxl n'est pas installé.\n\n"
                "Installez-le avec :\n  pip install openpyxl",
            )
            return

        if not self.gestionnaire.projets:
            messagebox.showinfo("Export", "Aucun projet à exporter.")
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichier Excel", "*.xlsx")],
            title="Enregistrer l'export Excel",
            initialfile="projets_export.xlsx",
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projets"

        # En-tête
        header_fill = PatternFill(
            start_color="1A1A3E", end_color="1A1A3E", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)

        ws.append([label for _, label, _ in COLUMNS])
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20

        # Données
        for projet in self.gestionnaire.projets:
            ws.append([self._val(projet, col_id) for col_id, _, _ in COLUMNS])

        # Largeur colonnes
        for idx, (_, _, px_w) in enumerate(COLUMNS, start=1):
            col_letter = openpyxl.utils.get_column_letter(idx)
            ws.column_dimensions[col_letter].width = max(px_w // 7, 10)

        wb.save(path)
        messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{path}")
